import openpyxl
import os


# 获取项目路径
class ExcelUtil:
    def get_object_path(self):
        return os.path.abspath(os.path.dirname(__file__)).split("common")[0]

    def read_excel(self):
        wb = openpyxl.load_workbook(self.get_object_path() + "data/login_data.xlsx")
        # 获得sheet对象
        sheet = wb['login']
        # 获取excel行和列
        # print(sheet.max_row, sheet.max_column)
        all_list = []
        for rows in range(2, sheet.max_row + 1):
            temp_list = []
            for columns in range(1, sheet.max_column + 1):
                temp_list.append(sheet.cell(rows, columns).value)
            all_list.append(temp_list)
        # print(all_list)
            # 返回一个列表，列表中有每行的数据
            # print(all_list)
        return all_list


# if __name__ == '__main__':
#     ExcelUtil().read_excel()
